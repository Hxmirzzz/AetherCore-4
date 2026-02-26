"""
Configuración de estilos para archivos Excel.

Separa la responsabilidad de definir estilos visuales de la lógica de negocio.
"""
from dataclasses import dataclass
from typing import Dict, Any
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

class ColorPalette:
    """Paleta de colores para los estilos"""
    HEADER_BG = "1F4E79"  # Azul oscuro
    HEADER_TEXT = "FFFFFF"  # Blanco
    
    LIGHT_ORANGE = "FFE0B2"  # Naranja claro
    LIGHT_BLUE = "DEEBF7"    # Azul claro
    LIGHT_GRAY = "D9D9D9"    # Gris claro
    LIGHT_BLUE_TOTAL = "D9E1F2"  # Azul claro para totales
    
    # Colores para encabezados de grupo (XML)
    GROUP_HEADER_MAIN = "4472C4"      # Azul principal
    GROUP_HEADER_INFO = "4F81BD"      # Azul info
    GROUP_HEADER_DENOM = "C0504D"     # Rojo denominaciones
    GROUP_HEADER_TOTAL = "9BBB59"     # Verde total
    
    # Color de texto
    TEXT_DARK_BLUE = "1F4E79"  # Azul oscuro para texto


class ExcelStyleConfig:
    """
    Configuración centralizada de estilos para Excel.
    
    Proporciona métodos para obtener estilos predefinidos consistentes.
    """
    
    @staticmethod
    def get_header_fill() -> PatternFill:
        """Relleno para encabezados de tabla"""
        return PatternFill(
            start_color=ColorPalette.HEADER_BG,
            end_color=ColorPalette.HEADER_BG,
            fill_type="solid"
        )
    
    @staticmethod
    def get_header_font() -> Font:
        """Fuente para encabezados de tabla"""
        return Font(
            color=ColorPalette.HEADER_TEXT,
            bold=True
        )
    
    @staticmethod
    def get_thin_border() -> Border:
        """Borde fino para celdas"""
        side = Side(style='thin')
        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )
    
    @staticmethod
    def get_medium_border_top_bottom() -> Border:
        """Borde medio solo arriba y abajo (para totales)"""
        return Border(
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
    
    @staticmethod
    def get_light_orange_fill() -> PatternFill:
        """Relleno naranja claro para filas alternas"""
        return PatternFill(
            start_color=ColorPalette.LIGHT_ORANGE,
            end_color=ColorPalette.LIGHT_ORANGE,
            fill_type="solid"
        )
    
    @staticmethod
    def get_light_blue_fill() -> PatternFill:
        """Relleno azul claro para filas alternas"""
        return PatternFill(
            start_color=ColorPalette.LIGHT_BLUE,
            end_color=ColorPalette.LIGHT_BLUE,
            fill_type="solid"
        )
    
    @staticmethod
    def get_light_gray_fill() -> PatternFill:
        """Relleno gris claro para secciones"""
        return PatternFill(
            start_color=ColorPalette.LIGHT_GRAY,
            end_color=ColorPalette.LIGHT_GRAY,
            fill_type="solid"
        )
    
    @staticmethod
    def get_light_blue_total_fill() -> PatternFill:
        """Relleno azul claro para celdas de totales"""
        return PatternFill(
            start_color=ColorPalette.LIGHT_BLUE_TOTAL,
            end_color=ColorPalette.LIGHT_BLUE_TOTAL,
            fill_type="solid"
        )
    
    @staticmethod
    def get_cell_alignment() -> Alignment:
        """Alineación estándar para celdas de datos"""
        return Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=False
        )
    
    @staticmethod
    def get_center_alignment() -> Alignment:
        """Alineación centrada para encabezados"""
        return Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
    
    @staticmethod
    def get_right_alignment() -> Alignment:
        """Alineación a la derecha (para etiquetas de totales)"""
        return Alignment(
            horizontal='right',
            vertical='center'
        )
    
    @staticmethod
    def get_normal_font() -> Font:
        """Fuente normal para datos"""
        return Font(size=10)
    
    @staticmethod
    def get_bold_font(size: int = 11, color: str = "000000") -> Font:
        """Fuente en negrita"""
        return Font(bold=True, size=size, color=color)
    
    @staticmethod
    def get_section_title_font() -> Font:
        """Fuente para títulos de sección"""
        return Font(bold=True, size=11)
    
    @staticmethod
    def get_main_title_font() -> Font:
        """Fuente para título principal"""
        return Font(bold=True, size=14, color=ColorPalette.HEADER_TEXT)
    
    # --- Estilos para XML (encabezados de grupo) ---
    
    @staticmethod
    def get_xml_main_header_style() -> Dict[str, Any]:
        """Estilo para encabezado principal de servicio (XML)"""
        return {
            'fill': PatternFill(
                start_color=ColorPalette.GROUP_HEADER_MAIN,
                end_color=ColorPalette.GROUP_HEADER_MAIN,
                fill_type="solid"
            ),
            'font': Font(color="FFFFFF", bold=True, size=14),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
    
    @staticmethod
    def get_xml_info_header_style() -> Dict[str, Any]:
        """Estilo para encabezado de información de entrega (XML)"""
        return {
            'fill': PatternFill(
                start_color=ColorPalette.GROUP_HEADER_INFO,
                end_color=ColorPalette.GROUP_HEADER_INFO,
                fill_type="solid"
            ),
            'font': Font(color="FFFFFF", bold=True, size=11),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
    
    @staticmethod
    def get_xml_denom_header_style() -> Dict[str, Any]:
        """Estilo para encabezado de denominaciones (XML)"""
        return {
            'fill': PatternFill(
                start_color=ColorPalette.GROUP_HEADER_DENOM,
                end_color=ColorPalette.GROUP_HEADER_DENOM,
                fill_type="solid"
            ),
            'font': Font(color="FFFFFF", bold=True, size=11),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
    
    @staticmethod
    def get_xml_total_header_style() -> Dict[str, Any]:
        """Estilo para encabezado de total general (XML)"""
        return {
            'fill': PatternFill(
                start_color=ColorPalette.GROUP_HEADER_TOTAL,
                end_color=ColorPalette.GROUP_HEADER_TOTAL,
                fill_type="solid"
            ),
            'font': Font(color="FFFFFF", bold=True, size=11),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
    
    @classmethod
    def get_all_styles_dict(cls) -> Dict[str, Any]:
        """
        Retorna un diccionario con todos los estilos (compatibilidad con código antiguo).
        
        DEPRECATED: Usar métodos individuales en su lugar.
        """
        return {
            'HEADER_FILL': cls.get_header_fill(),
            'HEADER_FONT': cls.get_header_font(),
            'THIN_BORDER': cls.get_thin_border(),
            'LIGHT_ORANGE_FILL': cls.get_light_orange_fill(),
            'LIGHT_BLUE_FILL': cls.get_light_blue_fill(),
            'CELL_ALIGNMENT': cls.get_cell_alignment(),
            'NORMAL_FONT': cls.get_normal_font()
        }