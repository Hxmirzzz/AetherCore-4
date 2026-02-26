"""
ExcelStyler centralizado para dar EXACTO el mismo estilo que el sistema anterior.
Mantén esta API estable: aplicar_estilos_excel(sheet, rowcount, startrow=2, codigo_point_index=None)
"""
from __future__ import annotations
from typing import Optional, Any
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=False)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

LIGHT_ORANGE_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
LIGHT_BLUE_FILL   = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

class ExcelStyler:
    """
    Estilo idéntico al antiguo:
      - encabezado con azul oscuro + texto blanco
      - filas alternas (naranja/azul) por grupo de CODIGO (opcional)
      - bordes finos y autoajuste de columnas
    API estable:
      aplicar_estilos_excel(worksheet, data_rows, startrow=0, codigo_point_index=None)
    """
    @staticmethod
    def aplicar_estilos_excel(
        worksheet: Any,
        data_rows: int,
        startrow: int = 0,
        codigo_point_index: Optional[int] = None
    ) -> None:
        try:
            if data_rows <= 0:
                logger.debug(f"[ExcelStyler] Sin filas para estilizar en '{worksheet.title}'.")
                return

            # Detectar la última columna con valor en la fila de encabezados
            max_col_with_value = 0
            for cell in worksheet[startrow + 1]:
                if cell.value is not None:
                    max_col_with_value = cell.column

            if max_col_with_value == 0:
                for r in range(startrow + 2, startrow + data_rows + 2):
                    for c in range(1, worksheet.max_column + 1):
                        if worksheet.cell(row=r, column=c).value is not None:
                            max_col_with_value = max(max_col_with_value, c)

            if max_col_with_value == 0:
                logger.warning(f"[ExcelStyler] No se detectaron columnas con datos en '{worksheet.title}'.")
                return

            # --- Encabezado ---
            header_row = startrow + 1  # openpyxl es 1-based
            for col in range(1, max_col_with_value + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER

            # --- Filas de datos (alternando por grupo de CODIGO si viene indice) ---
            current_codigo = None
            use_orange = True
            for r in range(header_row + 1, header_row + data_rows + 1):
                if codigo_point_index is not None:
                    codigo_cell = worksheet.cell(row=r, column=codigo_point_index + 1)
                    if codigo_cell.value != current_codigo:
                        current_codigo = codigo_cell.value
                        use_orange = not use_orange

                fill = LIGHT_ORANGE_FILL if use_orange else LIGHT_BLUE_FILL
                for c in range(1, max_col_with_value + 1):
                    cell = worksheet.cell(row=r, column=c)
                    cell.border = THIN_BORDER
                    # Centramos primeras columnas como el legado solía hacer (ajusta si quieres)
                    if c <= 10:
                        cell.alignment = CENTER
                    else:
                        cell.alignment = CELL_ALIGNMENT
                    cell.fill = fill

            # --- Auto ancho de columnas ---
            _col_widths = {}
            for r in range(header_row, header_row + data_rows + 1):
                for c in range(1, max_col_with_value + 1):
                    v = worksheet.cell(row=r, column=c).value
                    if v is None:
                        continue
                    w = len(str(v))
                    _col_widths[c] = max(_col_widths.get(c, 0), w)

            for c, w in _col_widths.items():
                adjusted = max(min(w + 3, 60), 15)
                worksheet.column_dimensions[get_column_letter(c)].width = adjusted

            logger.info(f"[ExcelStyler] Estilo aplicado a '{worksheet.title}' (rows={data_rows}, startrow={startrow}).")

        except Exception as e:
            logger.error(f"[ExcelStyler] Error al aplicar estilos: {e}", exc_info=True)