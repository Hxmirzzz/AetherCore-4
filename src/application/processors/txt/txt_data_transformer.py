from __future__ import annotations
from pathlib import Path
from typing import Dict, Any, List
import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment

from src.infrastructure.config.mapeos import TextosConstantes, DenominacionesConfig
from src.infrastructure.excel.excel_styler import ExcelStyler

logger = logging.getLogger(__name__)

_MAIN_HEADER_BG = "4472C4"
_INFO_BG        = "4F81BD"
_DENOM_BG       = "C0504D"
_TOTAL_BG       = "9BBB59"
_WHITE          = "FFFFFF"

class TxtDataTransformer:
    """
    - Normaliza columnas (opcional)
    - Calcula GENERAL si no existe (suma de columnas que comienzan con '$')
    - Crea XLSX con el MISMO layout/estilo que XML:
        F1 título, F2 bandas, F3 tabla, ExcelStyler, y “GRAN TOTAL” al pie
    """

    def to_dataframe(self, df_raw: pd.DataFrame) -> pd.DataFrame:
        df = df_raw.copy()

        # Si no trae GENERAL, súmalo de las columnas de denoms ($...)
        if "GENERAL" not in df.columns:
            denom_cols = [c for c in df.columns if isinstance(c, str) and c.startswith("$")]
            if denom_cols:
                def _parse_money(x: str) -> float:
                    if x is None: return 0.0
                    s = str(x).replace("$", "").replace(".", "").replace(",", "")
                    try:
                        return float(s) if s else 0.0
                    except Exception:
                        return 0.0
                df["GENERAL"] = df[denom_cols].applymap(_parse_money).sum(axis=1).astype(int)\
                    .map(lambda v: f"${v:,}".replace(",", "."))
        return df

    def write_excel_and_style(self, ruta_excel: Path, df: pd.DataFrame, titulo: str) -> bool:
        try:
            os.makedirs(ruta_excel.parent, exist_ok=True)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            ws = wb.create_sheet(title=titulo, index=0)
            self._agregar_titulo_y_grupo(ws, df, titulo)
            self._pegar_dataframe(ws, df, start_row=3)
            codigo_idx = df.columns.get_loc('CODIGO') if 'CODIGO' in df.columns else None
            ExcelStyler.aplicar_estilos_excel(ws, len(df), startrow=2, codigo_point_index=codigo_idx)
            self._agregar_gran_total(ws, df)

            wb.save(ruta_excel)
            logger.info("Excel TXT guardado: %s", ruta_excel.name)
            return True
        except Exception:
            logger.exception("Error creando/styling Excel TXT")
            return False

    # ===== Helpers (idénticos a los de XML para consistencia) =====
    def _pegar_dataframe(self, ws, df: pd.DataFrame, start_row: int) -> None:
        for r_idx, row_data in enumerate(
            dataframe_to_rows(df, index=False, header=True),
            start=start_row
        ):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    def _agregar_titulo_y_grupo(self, ws, df: pd.DataFrame, titulo: str) -> None:
        max_col = len(df.columns) if not df.empty else 1
        self._merge_and_style(ws, 1, 1, max_col, titulo, _MAIN_HEADER_BG, 14)
        if df.empty:
            return

        info_start, info_end, den_start, den_end, total_idx = self._find_ranges_for_groups(df)

        if info_start and info_end and info_end >= info_start:
            self._merge_and_style(ws, 2, info_start, info_end, TextosConstantes.ENCABEZADO_INFO_ENTREGA_XML, _INFO_BG, 12)
        if den_start and den_end and den_end >= den_start:
            self._merge_and_style(ws, 2, den_start, den_end, TextosConstantes.ENCABEZADO_DENOMINACIONES_XML, _DENOM_BG, 12)
        if total_idx:
            self._merge_and_style(ws, 2, total_idx, total_idx, TextosConstantes.ENCABEZADO_TOTAL_XML, _TOTAL_BG, 12)

    def _merge_and_style(self, ws, row, col_start, col_end, text, fill_hex, bold_size):
        if col_end < col_start:
            return
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start, value=text)
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell.font = Font(color=_WHITE, bold=True, size=bold_size)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _find_ranges_for_groups(self, df: pd.DataFrame):
        cols = list(df.columns)
        n = len(cols)
        first_denom = last_denom = total_idx = None

        for i, name in enumerate(cols, start=1):
            if isinstance(name, str) and name.startswith("$"):
                if first_denom is None:
                    first_denom = i
                last_denom = i
            if name == "GENERAL":
                total_idx = i

        if first_denom:
            info_start, info_end = 1, max(1, first_denom - 1)
        else:
            if total_idx:
                info_start, info_end = 1, max(1, total_idx - 1)
            else:
                info_start, info_end = 1, n

        den_start, den_end = (first_denom, last_denom) if first_denom and last_denom else (None, None)
        return info_start, info_end, den_start, den_end, total_idx

    def _agregar_gran_total(self, ws, df: pd.DataFrame) -> None:
        if df.empty:
            return
        last_row = ws.max_row + 1

        # Columna GENERAL a la derecha
        if "GENERAL" in df.columns:
            col_idx = list(df.columns).index("GENERAL") + 1
            lbl_cell = ws.cell(row=last_row, column=max(1, col_idx-1), value="GRAN TOTAL")
            lbl_cell.font = Font(bold=True)

            # Suma numérica (quitando formato)
            try:
                serie = (
                    df["GENERAL"].astype(str)
                    .str.replace("$", "", regex=False)
                    .str.replace(".", "", regex=False)
                    .str.replace(",", "", regex=False)
                )
                total_val = pd.to_numeric(serie, errors="coerce").sum(min_count=1)
            except Exception:
                total_val = None

            val_cell = ws.cell(row=last_row, column=col_idx,
                value=(f"${int(total_val):,}".replace(",", ".") if pd.notna(total_val) else ""))
            val_cell.font = Font(bold=True)
        else:
            ws.cell(row=last_row, column=1, value="GRAN TOTAL").font = Font(bold=True)
            
    def cel_consolidated(self,
        ruta_excel: Path,
        df_tipo1: pd.DataFrame | None,
        df_tipo2: pd.DataFrame | None,
        df_tipo3: pd.DataFrame | None,
        hoja_titulo: str = "Consolidado") -> bool:
        """
        Replica el layout del XML: título, sección INFO GENERAL (t1), TOTALES (t3),
        y DETALLE DE MOVIMIENTOS (t2) usando ExcelStyler.
        """
        
        try:
            os.makedirs(ruta_excel.parent, exist_ok=True)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            ws = wb.create_sheet(title=hoja_titulo, index=0)
            row = 1

            # Título
            ws.cell(row=row, column=1, value="CONSOLIDADO GENERAL")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            row += 2

            # ===== Sección: Información General (Tipo 1) =====
            if df_tipo1 is not None and not df_tipo1.empty:
                ws.cell(row=row, column=1, value="INFORMACIÓN GENERAL DEL ARCHIVO")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df_tipo1.columns))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                row += 1

                start_row_df1 = row
                for r_idx, row_data in enumerate(dataframe_to_rows(df_tipo1, index=False, header=True), start=start_row_df1):
                    for c_idx, value in enumerate(row_data, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    row = r_idx + 1

                from src.infrastructure.excel.excel_styler import ExcelStyler
                ExcelStyler.aplicar_estilos_excel(ws, len(df_tipo1), start_row_df1 - 1, None)
                row += 1

            # ===== Sección: Totales (Tipo 3) =====
            if df_tipo3 is not None and not df_tipo3.empty:
                ws.cell(row=row, column=1, value="TOTALES DEL ARCHIVO")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df_tipo3.columns))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                row += 1

                start_row_df3 = row
                for r_idx, row_data in enumerate(dataframe_to_rows(df_tipo3, index=False, header=True), start=start_row_df3):
                    for c_idx, value in enumerate(row_data, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    row = r_idx + 1

                from src.infrastructure.excel.excel_styler import ExcelStyler
                ExcelStyler.aplicar_estilos_excel(ws, len(df_tipo3), start_row_df3 - 1, None)
                row += 1

            # ===== Sección: Detalle de Movimientos (Tipo 2) =====
            if df_tipo2 is not None and not df_tipo2.empty:
                ws.cell(row=row, column=1, value="DETALLE DE MOVIMIENTOS")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df_tipo2.columns))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                row += 1

                df2_norm = self._normalize_df_tipo2_for_excel(df_tipo2)

                df2_export = df2_norm.drop('COD_SUC_INTERNO', axis=1, errors='ignore')

                start_row_df2 = row
                for r_idx, row_data in enumerate(dataframe_to_rows(df2_export, index=False, header=True), start=start_row_df2):
                    for c_idx, value in enumerate(row_data, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    row = r_idx + 1

                from src.infrastructure.excel.excel_styler import ExcelStyler
                codigo_punto_idx = df2_export.columns.get_loc('CODIGO PUNTO') if 'CODIGO PUNTO' in df2_export.columns else None
                ExcelStyler.aplicar_estilos_excel(ws, len(df2_export), start_row_df2 - 1, codigo_punto_idx)

            # Guardar
            wb.save(ruta_excel)
            return True
        except Exception:
            import logging
            logging.getLogger(__name__).exception("Error creando Excel consolidado TXT")
            return False
        
    def _normalize_df_tipo2_for_excel(self, df2: pd.DataFrame) -> pd.DataFrame:
        """
        Une filas de COP (códigos 1 y 2) en una sola, manteniendo layout y sumas:
        TOTAL_VALOR, CANT. BILLETE y columnas GAV* DENOMINACION/CANTIDAD.
        Otras divisas persisten separadas.
        """
        if df2 is None or df2.empty:
            return df2.copy()

        df = df2.copy()

        # 1) Estandarizar TIPO VALOR -> solo divisa (e.g. '1 - COP' / '2 - COP' -> 'COP')
        def _std_currency(x: str) -> str:
            s = str(x).strip()
            if " - " in s:
                return s.split(" - ", 1)[1].strip().upper()
            return s.upper()

        if 'TIPO VALOR' in df.columns:
            df['TIPO_VALOR_STD'] = df['TIPO VALOR'].astype(str).map(_std_currency)
        else:
            df['TIPO_VALOR_STD'] = 'COP'

        # 2) Columnas base para agrupar
        base_cols = [
            'CODIGO', 'FECHA SERVICIO', 'PRIORIDAD', 'CLIENTE', 'SERVICIO',
            'CODIGO PUNTO', 'NOMBRE PUNTO', 'CIUDAD', 'SUCURSAL', 'COD_SUC_INTERNO',
            'TIPO RUTA', 'TIPO PEDIDO'
        ]
        base_cols = [c for c in base_cols if c in df.columns]

        # 3) Columnas de gavetas
        gav_cols = [c for c in df.columns if isinstance(c, str) and c.startswith('GAV ')]
        denom_cols = [c for c in gav_cols if 'DENOMINACION' in c.upper()]
        cant_cols  = [c for c in gav_cols if 'CANTIDAD' in c.upper()]

        # 4) Preparar numéricos
        def _parse_money_to_int(v):
            if pd.isna(v): return 0
            s = str(v).replace('$','').replace('.','').replace(',','').strip()
            try: return int(float(s)) if s else 0
            except: return 0

        def _parse_int_str(v):
            if pd.isna(v): return 0
            s = str(v).replace('.','').replace(',','').strip()
            try: return int(float(s)) if s else 0
            except: return 0

        work = df.copy()
        work['_TOTAL_VALOR_NUM']   = work['TOTAL_VALOR'].map(_parse_money_to_int) if 'TOTAL_VALOR' in work.columns else 0
        work['_CANT_BILLETE_NUM']  = work['CANT. BILLETE'].map(_parse_int_str)   if 'CANT. BILLETE' in work.columns else 0
        for c in denom_cols:
            work[f'__NUM__{c}'] = work[c].map(_parse_money_to_int)
        for c in cant_cols:
            work[f'__NUM__{c}'] = work[c].map(_parse_int_str)

        # 5) Group key: base + TIPO_VALOR_STD (COP ya une 1 y 2)
        group_key = [*base_cols, 'TIPO_VALOR_STD']

        agg = {k: 'first' for k in group_key}
        if 'TIPO VALOR' in work.columns:
            agg['TIPO VALOR'] = 'first'
        agg['_TOTAL_VALOR_NUM']  = 'sum'
        agg['_CANT_BILLETE_NUM'] = 'sum'
        for c in denom_cols + cant_cols:
            agg[f'__NUM__{c}'] = 'sum'

        grouped = work.groupby(group_key, dropna=False, as_index=False).agg(agg)

        # 6) Reconstruir texto/formatos
        def _fmt_money(n): return f"${int(n):,}".replace(",", ".")
        def _fmt_int(n):    return f"{int(n):,}".replace(",", ".")

        if 'TOTAL_VALOR' in df.columns:
            grouped['TOTAL_VALOR'] = grouped['_TOTAL_VALOR_NUM'].map(_fmt_money)
        if 'CANT. BILLETE' in df.columns:
            grouped['CANT. BILLETE'] = grouped['_CANT_BILLETE_NUM'].map(_fmt_int)

        for c in denom_cols:
            grouped[c] = grouped[f'__NUM__{c}'].map(lambda n: _fmt_money(n) if n else "$0")
        for c in cant_cols:
            grouped[c] = grouped[f'__NUM__{c}'].map(_fmt_int)

        if 'TIPO VALOR' in grouped.columns:
            grouped['TIPO VALOR'] = grouped['TIPO_VALOR_STD']

        # 7) Orden de columnas (igual al original)
        base_order   = [c for c in df.columns if c in base_cols]
        totals_order = [c for c in ['TOTAL_VALOR', 'CANT. BILLETE'] if c in grouped.columns]

        def _gav_num(col):
            try:    return int(col.split('GAV ')[1].split(' ')[0])
            except: return 99999
        denom_order  = sorted([c for c in grouped.columns if c in denom_cols], key=_gav_num)
        cant_order   = sorted([c for c in grouped.columns if c in cant_cols],  key=_gav_num)

        extra_cols = [c for c in grouped.columns
            if c not in (base_order + totals_order + denom_order + cant_order
                + ['TIPO_VALOR_STD', '_TOTAL_VALOR_NUM', '_CANT_BILLETE_NUM'])
            and not c.startswith('__NUM__')]

        final_cols = base_order + totals_order + denom_order + cant_order + extra_cols
        grouped = grouped[[c for c in final_cols if c in grouped.columns]].copy()
        return grouped

    def write_excel_consolidated(
        self,
        ruta_excel: Path,
        df_tipo1: pd.DataFrame | None,
        df_tipo2: pd.DataFrame | None,
        df_tipo3: pd.DataFrame | None,
        hoja_titulo: str = "Consolidado"
    ) -> bool:
        """
        Genera un XLSX con layout idéntico al legado:
        - Título
        - Sección INFO (t1)
        - Sección TOTALES (t3)
        - Sección DETALLE (t2, normalizada: une COP 1/2)
        y aplica ExcelStyler para encabezados/filas/autoancho/alternado.
        """
        try:
            os.makedirs(ruta_excel.parent, exist_ok=True)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            ws = wb.create_sheet(title=hoja_titulo, index=0)
            row = 1

            # TÍTULO
            ws.cell(row=row, column=1, value="CONSOLIDADO GENERAL")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            row += 2

            # ===== t1: INFO GENERAL =====
            if df_tipo1 is not None and not df_tipo1.empty:
                ws.cell(row=row, column=1, value="INFORMACIÓN GENERAL DEL ARCHIVO")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df_tipo1.columns))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                row += 1

                start_row_df1 = row
                for r_idx, row_data in enumerate(dataframe_to_rows(df_tipo1, index=False, header=True), start=start_row_df1):
                    for c_idx, value in enumerate(row_data, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    row = r_idx + 1

                ExcelStyler.aplicar_estilos_excel(ws, len(df_tipo1), start_row_df1 - 1, None)
                row += 1

            # ===== t3: TOTALES =====
            if df_tipo3 is not None and not df_tipo3.empty:
                ws.cell(row=row, column=1, value="TOTALES DEL ARCHIVO")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df_tipo3.columns))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                row += 1

                start_row_df3 = row
                for r_idx, row_data in enumerate(dataframe_to_rows(df_tipo3, index=False, header=True), start=start_row_df3):
                    for c_idx, value in enumerate(row_data, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    row = r_idx + 1

                ExcelStyler.aplicar_estilos_excel(ws, len(df_tipo3), start_row_df3 - 1, None)
                row += 1

            # ===== t2: DETALLE =====
            if df_tipo2 is not None and not df_tipo2.empty:
                ws.cell(row=row, column=1, value="DETALLE DE MOVIMIENTOS")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df_tipo2.columns))
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                row += 1

                df2_export = df_tipo2.drop('COD_SUC_INTERNO', axis=1, errors='ignore')

                start_row_df2 = row
                for r_idx, row_data in enumerate(dataframe_to_rows(df2_export, index=False, header=True), start=start_row_df2):
                    for c_idx, value in enumerate(row_data, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    row = r_idx + 1

                codigo_punto_idx = df2_export.columns.get_loc('CODIGO PUNTO') if 'CODIGO PUNTO' in df2_export.columns else None
                ExcelStyler.aplicar_estilos_excel(ws, len(df2_export), start_row_df2 - 1, codigo_punto_idx)

            wb.save(ruta_excel)
            return True
        except Exception:
            logging.getLogger(__name__).exception("Error creando Excel consolidado TXT")
            return False