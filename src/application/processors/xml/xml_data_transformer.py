"""
Transformador de datos XML a DataFrames y creador de Excel (delegando estilos).
Replica columnas/diseño del procesador original:
    - Fila 1: Título mergeado + estilo (color principal)
    - Fila 2: Bandas de grupo (Info / Denominaciones / Total)
    - Fila 3+: Tabla (header + datos) -> estilizada por ExcelStyler.aplicar_estilos_excel
"""
from __future__ import annotations

import os
import logging
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from src.infrastructure.config.mapeos import TextosConstantes, DenominacionesConfig
from src.infrastructure.excel.excel_styler import ExcelStyler

logger = logging.getLogger(__name__)

_MAIN_HEADER_BG = "4472C4"   # Azul título principal
_INFO_BG        = "4F81BD"   # Azul cabecera "Información de entrega"
_DENOM_BG       = "C0504D"   # Rojo cabecera "Denominaciones"
_TOTAL_BG       = "9BBB59"   # Verde cabecera "Valores / Total"
_WHITE          = "FFFFFF"

class XmlDataTransformer:
    """
    Transforma datos XML a DataFrames y genera Excel con formato idéntico al original.
    """
    def to_dataframes(self, ordenes_filas: List[Dict[str, Any]], remesas_filas: List[Dict[str, Any]]) -> Dict[str, pd.DataFrame]:
        df_ordenes = self._create_ordered_dataframe(ordenes_filas) if ordenes_filas else pd.DataFrame()
        df_remesas = self._create_ordered_dataframe(remesas_filas) if remesas_filas else pd.DataFrame()
        return {"ordenes": df_ordenes, "remesas": df_remesas}
    
    def _create_ordered_dataframe(self, filas: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Crea DataFrame con orden específico de columnas según el original.
        
        Orden:
        1. Columnas de información
        2. Columnas de denominaciones (ordenadas según DenominacionesConfig.DENOMINACIONES)
        3. Columna GENERAL
        """
        if not filas:
            return pd.DataFrame()
        
        df = pd.DataFrame(filas)
        columnas_info = [
            'ID', 'FECHA DE ENTREGA', 'RANGO', 'ENTIDAD', 'CODIGO', 
            'NOMBRE PUNTO', 'TIPO DE SERVICIO', 'TRANSPORTADORA', 'CIUDAD'
        ]
        
        columnas_denom = []
        for denom_key in DenominacionesConfig.DENOMINACIONES:
            suf = denom_key[-2:]
            if suf in ("AD", "NF"):
                col_name = f'${denom_key[:-2]} {suf}'
            else:
                col_name = f'${denom_key}'
            columnas_denom.append(col_name)
            
        columnas_ordenadas = columnas_info + columnas_denom + ['GENERAL']
        columnas_finales = [col for col in columnas_ordenadas if col in df.columns]
        
        df = df[columnas_finales]
        
        return df
        
    def write_excel_and_style(self, ruta_excel: Path, df_ordenes: pd.DataFrame, df_remesas: pd.DataFrame) -> bool:
        """
        Escribe el Excel con:
            - Hoja PROVISIÓN si hay df_ordenes
            - Hoja RECOLECCIÓN si hay df_remesas
        Cada hoja:
            - F1 título mergeado
            - F2 bandas de grupo
            - F3+ tabla pegada con DataFrame
            - Estilos de tabla aplicados por ExcelStyler.aplicar_estilos_excel (startrow=2)
            - “GRAN TOTAL” al final (con suma en GENERAL si existe)
        """
        try:
            os.makedirs(ruta_excel.parent, exist_ok=True)
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
                
            # ===== PROVISIÓN =====
            if not df_ordenes.empty:
                ws_prov = wb.create_sheet(title=TextosConstantes.HOJA_PROVISION_XML, index=0)
                self._agregar_titulo_y_grupo(ws_prov, df_ordenes, TextosConstantes.SERVICIO_PROVISION_XML)
                self._pegar_dataframe(ws_prov, df_ordenes, start_row=3)
                
                codigo_idx = df_ordenes.columns.get_loc('CODIGO') if 'CODIGO' in df_ordenes.columns else None
                ExcelStyler.aplicar_estilos_excel(ws_prov, len(df_ordenes), startrow=2, codigo_point_index=codigo_idx)
                
                self._agregar_gran_total(ws_prov, df_ordenes, start_data_row=3)
                logger.info("Hoja '%s' creada con %d registros", TextosConstantes.HOJA_PROVISION_XML, len(df_ordenes))

            # ===== RECOLECCIÓN =====
            if not df_remesas.empty:
                idx = 0 if df_ordenes.empty else 1
                ws_reco = wb.create_sheet(title=TextosConstantes.HOJA_RECOLECCION_XML, index=idx)
                self._agregar_titulo_y_grupo(ws_reco, df_remesas, TextosConstantes.SERVICIO_RECOLECCION_XML)
                self._pegar_dataframe(ws_reco, df_remesas, start_row=3)
                
                codigo_idx = df_remesas.columns.get_loc('CODIGO') if 'CODIGO' in df_remesas.columns else None
                ExcelStyler.aplicar_estilos_excel(ws_reco, len(df_remesas), startrow=2, codigo_point_index=codigo_idx)
                
                self._agregar_gran_total(ws_reco, df_remesas, start_data_row=3)
                logger.info("Hoja '%s' creada con %d registros", TextosConstantes.HOJA_RECOLECCION_XML, len(df_remesas))

            wb.save(ruta_excel)
            logger.info("Excel guardado: %s", ruta_excel.name)
            return True
        except Exception:
            logger.exception("Error creando/styling Excel XML")
            return False

    # ===================== Helpers privados =====================
    def _pegar_dataframe(self, ws, df: pd.DataFrame, start_row: int) -> None:
        """
        Pega el DataFrame en la hoja a partir de start_row
        (incluye header=True para que la fila de encabezados quede en start_row).
        """
        for r_idx, row_data in enumerate(
            dataframe_to_rows(df, index=False, header=True),
            start=start_row
        ):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # ---------- Layout superior (título + bandas) ----------

    def _agregar_titulo_y_grupo(self, ws, df: pd.DataFrame, titulo: str) -> None:
        """
        Fila 1: título mergeado & estilizado.
        Fila 2: bandas de grupo auto-detectadas (Info / Denominaciones / Total).
        """
        # Título (fila 1)
        max_col = self._max_col_from_df(df) if not df.empty else 1
        self._merge_and_style(
            ws, row=1, col_start=1, col_end=max_col,
            text=titulo, fill_hex=_MAIN_HEADER_BG, bold_size=14
        )

        if df.empty:
            return

        # Bandas (fila 2)
        info_start, info_end, den_start, den_end, total_idx = self._find_ranges_for_groups(df)

        if info_start and info_end and info_end >= info_start:
            self._merge_and_style(
                ws, row=2, col_start=info_start, col_end=info_end,
                text=TextosConstantes.ENCABEZADO_INFO_ENTREGA_XML,
                fill_hex=_INFO_BG, bold_size=12
            )

        if den_start and den_end and den_end >= den_start:
            self._merge_and_style(
                ws, row=2, col_start=den_start, col_end=den_end,
                text=TextosConstantes.ENCABEZADO_DENOMINACIONES_XML,
                fill_hex=_DENOM_BG, bold_size=12
            )

        if total_idx:
            self._merge_and_style(
                ws, row=2, col_start=total_idx, col_end=total_idx,
                text=TextosConstantes.ENCABEZADO_TOTAL_XML,
                fill_hex=_TOTAL_BG, bold_size=12
            )

    def _merge_and_style(
        self,
        ws,
        row: int,
        col_start: int,
        col_end: int,
        text: str,
        fill_hex: str,
        bold_size: int = 12
    ) -> None:
        """Crea celda mergeada y aplica estilo consistente (color, fuente blanca, centrado)."""
        if col_end < col_start:
            return
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start, value=text)
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell.font = Font(color=_WHITE, bold=True, size=bold_size)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _find_ranges_for_groups(self, df: pd.DataFrame):
        """
        Detecta rangos para bandas de grupo:
            - info: desde columna 1 hasta antes de la primera columna con '$'
            - denoms: todas las columnas que empiezan con '$' (excepto GENERAL)
            - total: la columna 'GENERAL'
        
        Returns:
            Tupla (info_start, info_end, den_start, den_end, total_idx)
            Índices 1-based para openpyxl.
        """
        cols = list(df.columns)
        n = len(cols)

        first_denom = None
        last_denom = None
        total_idx = None

        for i, name in enumerate(cols, start=1):
            if isinstance(name, str) and name.startswith("$") and name != "GENERAL":
                if first_denom is None:
                    first_denom = i
                last_denom = i
            if name == "GENERAL":
                total_idx = i

        # Info: 1 hasta antes de la primera denom
        if first_denom:
            info_start, info_end = 1, max(1, first_denom - 1)
        else:
            if total_idx:
                info_start, info_end = 1, max(1, total_idx - 1)
            else:
                info_start, info_end = 1, n

        if first_denom and last_denom:
            den_start, den_end = first_denom, last_denom
        else:
            den_start, den_end = None, None

        return info_start, info_end, den_start, den_end, total_idx

    def _max_col_from_df(self, df: pd.DataFrame) -> int:
        return len(df.columns)

    # ---------- Gran Total ----------
    def _agregar_gran_total(self, ws, df: pd.DataFrame, start_data_row: int) -> None:
        """
        Inserta "GRAN TOTAL" con estilos idénticos al código original:
        - Etiqueta "TOTAL:" en la columna anterior a GENERAL (alineada a la derecha, azul oscuro)
        - Valor en la columna GENERAL (suma formateada, negrita, fondo celeste, bordes)
        
        Args:
            ws: Worksheet de openpyxl
            df: DataFrame con los datos
            start_data_row: Fila donde empiezan los datos (normalmente 3)
        """
        if df.empty:
            return
        
        total_row_idx = start_data_row + len(df) + 1
        
        if "GENERAL" not in df.columns:
            ws.cell(row=total_row_idx, column=1, value="GRAN TOTAL").font = Font(bold=True)
            return
        
        try:
            serie = (
                df["GENERAL"].astype(str)
                .str.replace("$", "", regex=False)
                .str.replace(".", "", regex=False)
                .str.replace(",", "", regex=False)
            )
            total_val = pd.to_numeric(serie, errors="coerce").sum(min_count=1)
        except Exception:
            total_val = None
            
        general_col_idx = list(df.columns).index("GENERAL") + 1
        text_col_idx = max(1, general_col_idx - 1)
        
        cell_text = ws.cell(row=total_row_idx, column=text_col_idx, value="GRAN TOTAL")
        cell_text.font = Font(bold=True, color="1F4E79")
        cell_text.alignment = Alignment(horizontal='right')
        
        cell_value = ws.cell(
            row=total_row_idx,
            column=general_col_idx,
            value=f"${int(total_val):,}".replace(",", ".") if pd.notna(total_val) else ""
        )
        cell_value.font = Font(bold=True, size=11, color="000000")
        cell_value.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell_value.border = Border(
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        if pd.notna(total_val):
            logger.info("Gran Total calculado: $%s", f"{int(total_val):,}".replace(",", "."))