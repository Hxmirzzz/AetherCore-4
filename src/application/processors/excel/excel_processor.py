from __future__ import annotations
from pathlib import Path
import logging
import os
import shutil
from datetime import datetime
import openpyxl
import uuid
import json

from pandas.core.generic import dt

from src.infrastructure.file_system.path_manager import PathManager
from src.application.processors.excel.excel_file_reader import ExcelFileReader
from src.application.processors.excel.excel_processor_factory import ExcelProcessorFactory
from src.presentation.api.internal_api_client import ApiService
from src.domain.value_objects.cliente_folder import ClienteFolder
from src.domain.value_objects.codigo_punto import CodigoPunto
from src.presentation.api.external_api_client import ExternalApiClient

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    Procesador principal de archivos Excel de solicitudes.
    """

    def __init__(
        self,
        reader: ExcelFileReader | None = None,
        api_service: ApiService | None = None,
        external_api: ExternalApiClient | None = None,
        path_manager: PathManager | None = None
    ):
        """
        Inicializa el procesador.
        """
        self._reader = reader or ExcelFileReader()
        self._api_service = api_service
        self._external_api = external_api
        self._path_manager = path_manager or PathManager()

    def procesar_archivo_excel(
        self,
        ruta_excel: Path,
        cliente_folder: ClienteFolder
    ) -> bool:
        """
        Procesa un archivo Excel de solicitudes.
        
        Args:
            ruta_excel: Path al archivo Excel
            cliente_folder: ClienteFolder del cliente
            
        Returns:
            True si procesó exitosamente, False en caso contrario
        """
        log_id = None
        try:
            cliente_name = cliente_folder.folder_name
            self._path_manager.create_request_structure(cliente_name)

            logger.info(f"== Procesando: {ruta_excel.name} para Cliente: {cliente_name} ==")

            if self._api_service:
                log_payload = {
                    "app": "AE_CORE_4",
                    "name": ruta_excel.name,
                    "fileType": "XLSX",
                    "estado": "PROCESANDO",
                    "recordCount": 0,
                    "payloadJson": None,
                    "correlationId": str(uuid.uuid4())
                }
                try:
                    res_log = self._api_service.register_event(log_payload)
                    if res_log:
                        log_id = res_log.get("id") or res_log.get("Id")
                except Exception as e:
                    logger.error(f"Error al registrar evento de inicio: {e}")
        
            info = self._reader.read_multiple_sheets(ruta_excel)
            if not info:
                self._actualizar_log_fallido(log_id, str(e), ruta_excel)
                self._manejar_excel_fallido(ruta_excel, cliente_name, "Archivo vacío o ilegible")
                return False
            
            try:
                mapper = ExcelProcessorFactory.get_mapper(cliente_folder.cod_cliente)
            except ValueError as e:
                self._actualizar_log_fallido(log_id, str(e), ruta_excel)
                self._manejar_excel_fallido(ruta_excel, cliente_name, str(e))
                return False

            nombre_hoja_params = next((k for k in info.keys() if "PARAMETRO" in k.upper()), None)
            if nombre_hoja_params and hasattr(mapper, 'actualizar_parametros'):
                mapper.actualizar_parametros(info.pop(nombre_hoja_params))

            dtos_a_enviar = []
            mapeo_filas_origen = {}
            
            for nombre_hoja, df in info.items():
                valido, error_msg = mapper.validar_estructura(df)
                if not valido:
                    logger.warning(f"Saltando hoja '{nombre_hoja}' - {error_msg}")
                    continue

                datos_hoja = mapper.mapear_a_dtos(df, f"{ruta_excel.name} [{nombre_hoja}]")
            
                for dto, idx_fila in datos_hoja:
                    punto = dto.cod_punto_origen if dto.cod_punto_origen != "FONDO" else dto.cod_punto_destino
                    punto_limpio = CodigoPunto.from_raw(str(punto)).parte_numerica.strip() 

                    dto.cod_punto_origen = punto_limpio
                    dto.cod_punto_destino = ""

                    dtos_a_enviar.append(dto)
                    key = len(dtos_a_enviar) - 1
                    mapeo_filas_origen[key] = (nombre_hoja, idx_fila, dto.numero_pedido)
                
            if not dtos_a_enviar:
                self._actualizar_log_fallido(log_id, "No se encontraron registros válidos en el archivo", ruta_excel)
                self._manejar_excel_fallido(ruta_excel, cliente_name, "No se encontraron registros válidos en el archivo")
                return False

            cantidad_registros = len(dtos_a_enviar)
            payload_str = json.dumps([dto.__dict__ for dto in dtos_a_enviar], default=str)
            logger.info(f"🚀 Enviando {cantidad_registros} servicios a VCashApp vía API...")
            respuesta = self._api_service.upload_services(dtos_a_enviar)

            if respuesta:
                if log_id and self._api_service:
                    try:
                        self._api_service.update_event(log_id, {
                            "estado": "COMPLETADO",
                            "responseJson": json.dumps(respuesta),
                            "errorDetails": None,
                            "processedBy": "AE4",
                            "filePath": str(ruta_excel.absolute()),
                            "recordCount": cantidad_registros,
                            "payloadJson": payload_str
                        })
                    except Exception as e:
                        logger.error(f"Error al actualizar log: {e}")
                if self._external_api:
                    try:
                        payload_externo = self._preparar_payload_externo(dtos_a_enviar)

                        logger.info(f"🌐 Disparando {len(payload_externo)} servicios a la API Externa en modo Bulk...")

                        res_ext = self._external_api.create_bulk_orders(payload_externo)
                        
                        if res_ext.get("status") == "success":
                            logger.info("✅ Servicios enviados exitosamente a la API Externa")
                        else:
                            logger.warning(f"⚠️ Algunos servicios fallaron en la API Externa: {res_ext.get('message', 'Unknown error')}")
                    except Exception as e:
                        logger.error(f"Error al enviar servicios a la API Externa: {e}")

                return self._gestionar_finalizacion(ruta_excel, cliente_name, respuesta, mapeo_filas_origen)
            else:
                self._actualizar_log_fallido(log_id, "Error al enviar servicios a VCashApp", ruta_excel)
                self._manejar_excel_fallido(ruta_excel, cliente_name, "Error al enviar servicios a VCashApp")
                return False
            
        except Exception as e:
            logger.exception(f"❌ Error crítico: {e}")
            self._actualizar_log_fallido(log_id, f"Error crítico: {str(e)}", ruta_excel)
            self._manejar_excel_fallido(ruta_excel, cliente_name, str(e))
            return False

    def _gestionar_finalizacion(self, ruta_excel: Path, cliente_name: str, respuesta: dict, mapeo: dict) -> bool:
        """
        Gestiona la finalización del procesamiento del archivo Excel.
        """
        exitosos_indices = {}
        fallidos_indices = {}
        errores_log = []
        
        for idx, result in enumerate(respuesta.get('details', [])):
            hoja, fila_idx, pedido = mapeo[idx]

            if result['success']:
                if hoja not in exitosos_indices: exitosos_indices[hoja] = set()
                exitosos_indices[hoja].add(fila_idx)
            else:
                if hoja not in fallidos_indices: fallidos_indices[hoja] = set()
                fallidos_indices[hoja].add(fila_idx)
                errores_log.append(f"Fila {fila_idx + 6} (Hoja {hoja}): Pedido {pedido} -> {result['message']}")

        total_exitosos = sum(len(v) for v in exitosos_indices.values())
        total_fallidos = sum(len(v) for v in fallidos_indices.values())

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_base = ruta_excel.stem.replace('_NOVEDADES', '').replace('_OK', '').split('_202')[0]

        if total_exitosos > 0:
            gestionados_dir =  self._path_manager.get_gestionado_path(cliente_name)
            gestionados_dir.mkdir(parents=True, exist_ok=True)
            ruta_ok = gestionados_dir / f"{nombre_base}_OK_{timestamp}{ruta_excel.suffix}"

            if total_fallidos == 1:
                shutil.move(str(ruta_excel), str(ruta_ok))
                logger.info(f"✅ Archivo completo movido a Gestionados")
            else:
                self._generar_copia_filtrada(ruta_excel, ruta_ok, exitosos_indices)
                logger.info(f"📁 Copia de exitosos guardada en GESTIONADOS")

        if total_fallidos > 0:
            novedades_dir = self._path_manager.get_novedades_path(cliente_name)
            novedades_dir.mkdir(parents=True, exist_ok=True)
            ruta_novedades = novedades_dir / f"{nombre_base}_NOVEDADES_{timestamp}{ruta_excel.suffix}"
            self._generar_copia_filtrada(ruta_excel, ruta_novedades, fallidos_indices, borrar_hojas_vacias=True)
            
            ruta_txt = novedades_dir / f"{nombre_base}_NOVEDADES_{timestamp}.txt"
            with open(ruta_txt, 'w', encoding='utf-8') as f:
                f.write(f"REPORTE DE NOVEDADES - AETHERCORE 4\n")
                f.write(f"Archivo: {ruta_excel.name}\n")
                f.write(f"Resultado API: {respuesta.get('summary')}\n")
                f.write("-" * 60 + "\n")
                for err in errores_log:
                    f.write(f"- {err}\n")
            
            logger.warning(f"⚠️ Se generó archivo de NOVEDADES con {total_fallidos} errores")


        if ruta_excel.exists():
            try:
                os.remove(ruta_excel)
            except Exception as e:
                logger.error(f"❌ Error al eliminar archivo original: {e}")  

        logger.info(f"=== PROCESO FINALIZADO: {total_exitosos} OK / {total_fallidos} ERROR ===")
        return True          

    def _generar_copia_filtrada(self, ruta_origen: Path, ruta_destino: Path, filas_a_mantener: dict, borrar_hojas_vacias: bool = True):
        """
        Genera una copia del archivo Excel filtrando solo las filas especificadas.
        """
        try:
            shutil.copy2(ruta_origen, ruta_destino)
            wb = openpyxl.load_workbook(ruta_destino)
            KEYWORDS = ExcelFileReader.KEYWORDS_HEADER

            hojas_a_borrar = []
            
            for sheet_name in wb.sheetnames:
                if "PARAMETRO" in sheet_name.upper():
                    continue

                indices_a_mantener = filas_a_mantener.get(sheet_name, set())
                ws = wb[sheet_name]
                
                if not indices_a_mantener:
                    if borrar_hojas_vacias and len(wb.sheetnames) > 1:
                        hojas_a_borrar.append(sheet_name)
                    else:
                        self._limpiar_datos_hoja(ws, KEYWORDS, set())
                    continue
                    
                self._limpiar_datos_hoja(ws, KEYWORDS, indices_a_mantener)
            
            for sh_name in hojas_a_borrar:
                if len(wb.sheetnames) > 1:
                    del wb[sh_name]
            
            wb.save(ruta_destino)
            wb.close()
            
        except Exception as e:
            logger.error(f"Error generando copia filtrada {ruta_destino.name}: {e}")

    def _limpiar_datos_hoja(self, ws, keywords: list, indices_a_mantener: set):
        """
        Limpia los datos de una hoja, manteniendo solo las filas especificadas.
        """
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), start=1):
            row_str = [str(c).upper().strip() if c else "" for c in row]
            if any(k in row_str for k in keywords):
                header_row_idx = i
                break
        
        if not header_row_idx: return

        start_data_row = header_row_idx + 1
        max_row = ws.max_row
        filas_a_mantener = {start_data_row + idx for idx in indices_a_mantener}

        for r in range(start_data_row, max_row + 1):
            if r not in filas_a_mantener:
                for cell in ws[r]:
                    cell.value = None

                ws.row_dimensions[r].hidden = True

    def _manejar_excel_fallido(self, ruta_excel: Path, cliente_name: str, razon_error: str):
        """
        Maneja archivos Excel que fallaron en el procesamiento.
        """
        logger.error(f"   Razón: {razon_error}")
        try:
            errores_dir = self._path_manager.get_errores_path(cliente_name)
            errores_dir.mkdir(exist_ok=True)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre = f"{ruta_excel.stem}_ERROR_{timestamp}{ruta_excel.suffix}"
            destino = errores_dir / nombre

            os.rename(ruta_excel, destino)

            log_path = destino.with_suffix('.txt')
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(f"Error: {razon_error}\nFecha: {datetime.now()}")
            
            logger.info(f"Log de error creado: {destino.name}")
        except Exception as e:
            logger.error(f"Error moviendo archivo fallido: {e}")

    def _actualizar_log_fallido(self, log_id: int | None, error_msg: str, ruta_excel: Path, record_count: int = 0, payload_str: str = None):
        """
        Actualiza el log de un archivo fallido.
        """
        if log_id and self._api_service:
            try:
                self._api_service.update_event(log_id, {
                    "estado": "FALLIDO",
                    "responseJson": None,
                    "errorDetails": error_msg,
                    "processedBy": "AE4",
                    "filePath": str(ruta_excel.absolute()),
                    "recordCount": record_count,
                    "payloadJson": payload_str
                })
            except Exception as e:
                logger.error(f"Error al actualizar log fallido {log_id}: {e}")

    def _preparar_payload_externo(self, dtos: list) -> list:
        """
        Prepara el payload para enviar a la API externa.
        """
        payload = []
        for dto in dtos:
            punto = str(dto.cod_punto_origen).strip()
            raw_value = dto.valor_total_declarado
            if not raw_value:
                raw_value = dto.valor_billete + dto.valor_moneda
            logger.info(f"🔍 DEBUG MONTO - Valor crudo: '{raw_value}' | Tipo: {type(raw_value)}")

            try:
                clean_value = str(raw_value).replace('$', '').replace(',', '').replace(' ', '').strip()
                total_service = str(int(float(clean_value)))
            except (ValueError, TypeError):
                logger.error(f"❌ ERROR CONVERSION - No se pudo convertir '{raw_value}' a entero")
                total_service = "0"

            service = {
                "client_code": str(dto.cod_cliente),
                "service_type": "SC",
                "service_date": str(dto.fecha_programacion),
                "time_window_start": "08:00:00.000Z",
                "time_window_end": "18:00:00.000Z",
                "declared_amount": total_service,
                "currency": "COP",
                "observations": str(dto.observaciones or ""),
                "bank_name": "",
                "bank_account_number": "",
                "bank_account_holder": "",
                "requested_denominations": []
            }

            if "-" in punto:
                service["atm_code"] = punto
            elif punto:
                service["point_code"] = punto

            payload.append(service)
        return payload